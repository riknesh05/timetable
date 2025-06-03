from flask import Flask, render_template, request, redirect, url_for, send_file
import json
import os
import io
from openpyxl import Workbook

app = Flask(__name__)

DATA_DIR = 'data'
STAFF_FILE = os.path.join(DATA_DIR, 'staff.json')
os.makedirs(DATA_DIR, exist_ok=True)

def load_staff():
    if not os.path.exists(STAFF_FILE):
        return []
    with open(STAFF_FILE, 'r') as f:
        return json.load(f)

def save_staff(data):
    with open(STAFF_FILE, 'w') as f:
        json.dump(data, f, indent=2)

def generate_staff_id(staff_list):
    if not staff_list:
        return 'S001'
    last_id = staff_list[-1]['id']
    num = int(last_id[1:]) + 1
    return f'S{num:03d}'

@app.route('/staff')
def staff_page():
    staff = load_staff()
    return render_template('staff.html', staff=staff)

@app.route('/staff/add', methods=['POST'])
def add_staff():
    name = request.form['name'].strip()
    subjects = request.form.getlist('subjects')
    departments = request.form.getlist('departments')

    staff_list = load_staff()
    new_id = generate_staff_id(staff_list)

    staff_list.append({
        'id': new_id,
        'name': name,
        'subjects': subjects,
        'departments': departments
    })

    save_staff(staff_list)
    return redirect(url_for('staff_page'))

@app.route('/staff/delete/<staff_id>')
def delete_staff(staff_id):
    staff_list = load_staff()
    staff_list = [s for s in staff_list if s['id'] != staff_id]
    save_staff(staff_list)
    return redirect(url_for('staff_page'))

@app.route('/timetable', methods=['GET', 'POST'])
def timetable_page():
    staff = load_staff()
    departments = ['CSE 2nd', 'ECE 1st', 'IT 4th']
    days_options = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    periods = 6

    timetable = {}

    if request.method == 'POST':
        selected_days = request.form.getlist('days')
        periods = int(request.form.get('periods'))

        # Initialize empty timetable
        for dept in departments:
            timetable[dept] = {}
            for day in selected_days:
                timetable[dept][day] = [''] * periods

        # Round-robin allocation for each department
        for dept in departments:
            relevant_staff = [s for s in staff if dept in s['departments']]

            # Gather all (subject, staff_id) pairs for the dept
            subjects_staff_pairs = []
            for s in relevant_staff:
                for subj in s['subjects']:
                    subjects_staff_pairs.append((subj, s['id']))

            total = len(subjects_staff_pairs)
            idx = 0

            for day in selected_days:
                for p in range(periods):
                    if total > 0:
                        subj, sid = subjects_staff_pairs[idx % total]
                        timetable[dept][day][p] = f"{subj} ({sid})"
                        idx += 1
                    else:
                        timetable[dept][day][p] = "-"

    return render_template('timetable.html',
                           staff=staff,
                           timetable=timetable,
                           departments=departments,
                           days_options=days_options,
                           periods=periods)

@app.route('/download-timetable', methods=['POST'])
def download_timetable():
    data = request.form.get('timetable_json')
    timetable = json.loads(data)

    wb = Workbook()
    for dept, days in timetable.items():
        ws = wb.create_sheet(title=dept[:31])
        header = ['Day'] + [f'Period {i+1}' for i in range(len(next(iter(days.values()))))]
        ws.append(header)
        for day, periods in days.items():
            ws.append([day] + periods)

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        std = wb['Sheet']
        wb.remove(std)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='timetable.xlsx')

if __name__ == '__main__':
    app.run(debug=True)
